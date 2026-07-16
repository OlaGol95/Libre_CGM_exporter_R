from __future__ import annotations

import argparse
import csv
import getpass
from io import BytesIO
import math
import os
import re
import shutil
import subprocess
import sys
import unicodedata
from datetime import date, datetime, time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
DEFAULT_LIBRE_DIR = Path(os.environ.get("LIBRE_DIR", str(APP_DIR))).expanduser()
DATE_FILE_NAME = "data_visits.xlsx"
DATE_ARCHIVE_NAME = "data_visits_protected.zip"
PASSWORD_ENV = "LIBRE_VISIT_DATES_PASSWORD"
PASSWORD_FILE = APP_DIR / "libre_dates_password.txt"
VISITS = ("V1", "V3", "V4", "V6")


def normalize_text(value):
    text = str(value or "").strip().lower().replace("\ufeff", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch for ch in text if ch.isalnum())


def patient_folder_name(patient_id):
    if pd.isna(patient_id):
        return ""
    text = str(patient_id).strip()
    if text.upper().startswith("G"):
        digits = "".join(ch for ch in text if ch.isdigit())
        return f"G{int(digits):04d}" if digits else text.upper()
    try:
        return f"G{int(float(text)):04d}"
    except ValueError:
        digits = "".join(ch for ch in text if ch.isdigit())
        return f"G{int(digits):04d}" if digits else text


def parse_excel_datetime(value):
    if value is None or value == "" or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time())
    text = str(value).strip()
    if not text or text in {"-", "nan", "NaT"}:
        return None
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    return None if pd.isna(parsed) else parsed.to_pydatetime()


def range_end_for_coverage(value):
    if value is None:
        return None
    if value.time() == time(0, 0, 0):
        return datetime.combine(value.date(), time(23, 59, 59))
    return value


def detect_separator(sample_text):
    try:
        return csv.Sniffer().sniff(sample_text, delimiters=",;\t").delimiter
    except Exception:
        return None


def try_read_csv(path):
    encodings = ("utf-8-sig", "utf-8", "cp1250", "latin1")
    header_rows = (2, 0, 1, 3)
    errors = []
    for encoding in encodings:
        try:
            sample = path.read_text(encoding=encoding, errors="replace")[:8192]
        except Exception as exc:
            errors.append(f"{encoding}: {exc}")
            continue
        separators = []
        detected = detect_separator(sample)
        if detected:
            separators.append(detected)
        for sep in (",", ";", "\t"):
            if sep not in separators:
                separators.append(sep)
        for sep in separators:
            for header_row in header_rows:
                try:
                    df = pd.read_csv(path, sep=sep, header=header_row, encoding=encoding, low_memory=False)
                    if df is not None and not df.empty and len(df.columns) >= 2:
                        return df
                except Exception as exc:
                    errors.append(f"{encoding}/{repr(sep)}/header={header_row}: {exc}")
    raise ValueError("Could not read Libre CSV:\n" + "\n".join(errors[:20]))


def find_timestamp_column(df):
    wanted = {
        normalize_text("Znacznik czasu w urządzeniu"),
        normalize_text("Znacznik czasu w urzadzeniu"),
        normalize_text("Device Timestamp"),
        normalize_text("Timestamp"),
    }
    for column in df.columns:
        if normalize_text(column) in wanted:
            return column
    for column in df.columns:
        normalized = normalize_text(column)
        if "znacznikczasu" in normalized or "timestamp" in normalized:
            return column
    if len(df.columns) >= 3:
        return df.columns[2]
    raise ValueError("Timestamp column not found.")


def glucose_column_priority(column_name):
    normalized = normalize_text(column_name)
    if ("gluko" not in normalized and "glucose" not in normalized) or "mgdl" not in normalized:
        return None
    if "keton" in normalized:
        return None
    if "history" in normalized or "historyczne" in normalized:
        return 0
    if "skan" in normalized or "scan" in normalized:
        return 1
    if "paska" in normalized or "strip" in normalized:
        return 2
    return 3


def find_glucose_columns(df):
    ranked = []
    for column in df.columns:
        priority = glucose_column_priority(column)
        if priority is not None:
            ranked.append((priority, column))
    ranked.sort(key=lambda item: item[0])
    return [column for _priority, column in ranked]


def glucose_value_from_row(row, glucose_columns):
    for column in glucose_columns:
        value = row.get(column)
        if pd.isna(value):
            continue
        text = str(value).strip().replace(",", ".")
        if not text:
            continue
        try:
            parsed = float(text)
        except ValueError:
            continue
        if not math.isnan(parsed):
            return parsed
    return None


def extract_last_numeric_value(row_values, skip_index):
    for idx in range(len(row_values) - 1, -1, -1):
        if idx == skip_index:
            continue
        value = row_values[idx]
        if pd.isna(value):
            continue
        text = str(value).strip().replace(",", ".")
        if not text:
            continue
        try:
            return float(text)
        except ValueError:
            continue
    return None


def extract_timestamp_and_glucose(df):
    timestamp_column = find_timestamp_column(df)
    timestamp_idx = list(df.columns).index(timestamp_column)
    glucose_columns = find_glucose_columns(df)
    parsed = pd.to_datetime(df[timestamp_column].astype(str).str.strip(), errors="coerce", dayfirst=True)
    if parsed.notna().sum() == 0:
        parsed = pd.to_datetime(df[timestamp_column].astype(str).str.strip(), errors="coerce", dayfirst=False)
    rows = []
    for row_index, timestamp in parsed.items():
        if pd.isna(timestamp):
            continue
        row = df.iloc[row_index]
        glucose = glucose_value_from_row(row, glucose_columns)
        if glucose is None and not glucose_columns:
            glucose = extract_last_numeric_value(row.tolist(), timestamp_idx)
        if glucose is None:
            continue
        rows.append({"timestamp": timestamp.to_pydatetime(), "glucose": glucose})
    return pd.DataFrame(rows).sort_values("timestamp").reset_index(drop=True) if rows else pd.DataFrame()


def visit_date_password():
    value = os.environ.get(PASSWORD_ENV, "")
    if value:
        return value
    if PASSWORD_FILE.exists():
        return PASSWORD_FILE.read_text(encoding="utf-8").strip()
    return getpass.getpass("Password for data_visits_protected.zip: ")


def require_pyzipper():
    try:
        import pyzipper  # type: ignore
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: pyzipper\n"
            "Install it with:\n"
            "    pip install pyzipper"
        ) from exc
    return pyzipper


def load_visit_workbook(libre_dir, no_password=False):
    archive_path = libre_dir / DATE_ARCHIVE_NAME
    excel_path = libre_dir / DATE_FILE_NAME
    if archive_path.exists() and not no_password:
        pyzipper = require_pyzipper()
        password = visit_date_password()
        if not password:
            raise SystemExit("Visit-date archive password is required.")
        try:
            with pyzipper.AESZipFile(archive_path) as archive:
                archive.setpassword(password.encode("utf-8"))
                with archive.open(DATE_FILE_NAME) as source:
                    payload = source.read()
        except RuntimeError as exc:
            raise SystemExit("Invalid password or unreadable visit-date archive.") from exc
        except KeyError as exc:
            raise SystemExit(f"Missing {DATE_FILE_NAME} inside {DATE_ARCHIVE_NAME}.") from exc
        return load_workbook(BytesIO(payload), data_only=True, read_only=True)
    if excel_path.exists():
        return load_workbook(excel_path, data_only=True, read_only=True)
    raise SystemExit(f"Missing {DATE_ARCHIVE_NAME}: {archive_path}")


def load_patients_from_workbook(workbook):
    try:
        sheet = workbook[workbook.sheetnames[0]]
        headers = [str(sheet.cell(row=1, column=c).value or f"COL{c}").strip() for c in range(1, sheet.max_column + 1)]
        rows = []
        for row_num in range(2, sheet.max_row + 1):
            values = {headers[c - 1]: sheet.cell(row=row_num, column=c).value for c in range(1, sheet.max_column + 1)}
            folder = patient_folder_name(values.get("ID"))
            if not folder:
                continue
            normalized_headers = {normalize_text(header): header for header in headers}
            visits = {}
            for visit in VISITS:
                start_key = normalized_headers.get(normalize_text(f"{visit}start"))
                end_key = normalized_headers.get(normalize_text(f"{visit}end"))
                visits[visit] = {
                    "start": parse_excel_datetime(values.get(start_key)) if start_key else None,
                    "end": parse_excel_datetime(values.get(end_key)) if end_key else None,
                }
            rows.append({"folder": folder, "visits": visits})
        return rows
    finally:
        workbook.close()


def load_patients(libre_dir, no_password=False):
    workbook = load_visit_workbook(libre_dir, no_password=no_password)
    return load_patients_from_workbook(workbook)


def patient_source_csvs(libre_dir, patient_folder):
    folder = libre_dir / patient_folder
    if not folder.exists():
        return []
    generated_prefixes = (f"{patient_folder}_CGM",)
    return [
        path for path in sorted(folder.glob("*.csv"), key=lambda p: p.name.lower())
        if not any(path.stem.startswith(prefix) for prefix in generated_prefixes)
    ]


def visit_frames_for_patient(libre_dir, patient):
    source_paths = patient_source_csvs(libre_dir, patient["folder"])
    columns = ["patient_id", "visit", "timestamp", "glucose", "source_file"]
    by_visit = {visit: pd.DataFrame(columns=columns) for visit in VISITS}
    if not source_paths:
        return by_visit
    all_rows = []
    for source_path in source_paths:
        frame = extract_timestamp_and_glucose(try_read_csv(source_path))
        if frame.empty:
            continue
        frame["patient_id"] = patient["folder"]
        frame["source_file"] = source_path.name
        all_rows.append(frame[["patient_id", "timestamp", "glucose", "source_file"]])
    if not all_rows:
        return by_visit
    combined = pd.concat(all_rows, ignore_index=True)
    combined = combined.dropna(subset=["timestamp"]).sort_values("timestamp")
    combined = combined.drop_duplicates(subset=["timestamp", "glucose", "source_file"])
    for visit in VISITS:
        start = patient["visits"][visit]["start"]
        end = range_end_for_coverage(patient["visits"][visit]["end"])
        if start is None or end is None:
            continue
        selected = combined[(combined["timestamp"] >= start) & (combined["timestamp"] <= end)].copy()
        if selected.empty:
            continue
        selected.insert(1, "visit", visit.lower())
        by_visit[visit] = selected[columns].reset_index(drop=True)
    return by_visit


def python_summary(patient, visit_frames):
    rows = []
    for visit in VISITS:
        frame = visit_frames.get(visit, pd.DataFrame())
        glucose = pd.to_numeric(frame.get("glucose"), errors="coerce").dropna() if not frame.empty else pd.Series(dtype=float)
        timestamps = pd.to_datetime(frame.get("timestamp"), errors="coerce").dropna() if not frame.empty else pd.Series(dtype="datetime64[ns]")
        rows.append({
            "patient_id": patient["folder"],
            "visit": visit.lower(),
            "n": int(glucose.shape[0]),
            "start": "" if timestamps.empty else timestamps.min(),
            "end": "" if timestamps.empty else timestamps.max(),
            "mean_glucose": None if glucose.empty else round(float(glucose.mean()), 3),
            "median_glucose": None if glucose.empty else round(float(glucose.median()), 3),
            "sd_glucose": None if len(glucose) < 2 else round(float(glucose.std(ddof=1)), 3),
            "min_glucose": None if glucose.empty else round(float(glucose.min()), 3),
            "max_glucose": None if glucose.empty else round(float(glucose.max()), 3),
            "pct_below_54": None if glucose.empty else round(float((glucose < 54).mean() * 100), 3),
            "pct_below_70": None if glucose.empty else round(float((glucose < 70).mean() * 100), 3),
            "pct_70_180": None if glucose.empty else round(float(((glucose >= 70) & (glucose <= 180)).mean() * 100), 3),
            "pct_above_180": None if glucose.empty else round(float((glucose > 180).mean() * 100), 3),
            "pct_above_250": None if glucose.empty else round(float((glucose > 250).mean() * 100), 3),
        })
    return pd.DataFrame(rows)


def find_rscript():
    found = shutil.which("Rscript")
    if found:
        return found
    r_root = Path(r"C:\Program Files\R")
    if r_root.exists():
        candidates = sorted(r_root.glob(r"R-*\bin\Rscript.exe"), reverse=True)
        if candidates:
            return str(candidates[0])
    return ""


def run_iglu_summary(visit_frames):
    rscript = find_rscript()
    visits_lower = [visit.lower() for visit in VISITS]
    if not rscript:
        return pd.DataFrame({"visit": visits_lower, "iglu_status": ["Rscript not found"] * len(VISITS)})
    rows = []
    for visit in VISITS:
        frame = visit_frames.get(visit, pd.DataFrame())
        for _idx, row in frame.iterrows():
            rows.append({
                "id": visit.lower(),
                "time": pd.to_datetime(row["timestamp"]).strftime("%Y-%m-%d %H:%M:%S"),
                "gl": row["glucose"],
            })
    if not rows:
        return pd.DataFrame({"visit": visits_lower, "iglu_status": ["no glucose data"] * len(VISITS)})
    stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    input_path = APP_DIR / f"_tmp_iglu_{stamp}_input.csv"
    output_path = APP_DIR / f"_tmp_iglu_{stamp}_output.csv"
    script_path = APP_DIR / f"_tmp_iglu_{stamp}.R"
    try:
        pd.DataFrame(rows).to_csv(input_path, index=False, encoding="utf-8")
        script_path.write_text("\n".join([
    "args <- commandArgs(trailingOnly = TRUE)",
    "options(repos = c(CRAN = 'https://cloud.r-project.org'))",
    "if (!requireNamespace('iglu', quietly = TRUE)) {",
    "  install.packages('iglu')",
    "}",
    "suppressPackageStartupMessages(library(iglu))",
            "d <- read.csv(args[1], stringsAsFactors = FALSE)",
            "d$time <- as.POSIXct(d$time, tz = 'UTC')",
            "d$gl <- as.numeric(d$gl)",
            "d <- d[!is.na(d$time) & !is.na(d$gl), ]",
            "d <- d[order(d$id, d$time), ]",
            "ids <- unique(d$id)",
            "prefix_metric_cols <- function(res, label) {",
            "  res <- as.data.frame(res)",
            "  if (!'id' %in% names(res)) res$id <- ids[seq_len(nrow(res))]",
            "  non_id <- setdiff(names(res), 'id')",
            "  if (length(non_id) == 1) names(res)[names(res) == non_id] <- label",
            "  else if (length(non_id) > 1) names(res)[names(res) %in% non_id] <- paste(label, non_id, sep = '_')",
            "  res",
            "}",
            "safe_metric <- function(label, expr) {",
            "  tryCatch(prefix_metric_cols(eval(expr), label), error = function(e) {",
            "    out <- data.frame(id = ids); out[[paste0(label, '_error')]] <- conditionMessage(e); out",
            "  })",
            "}",
            "roc_summary <- function(data) {",
            "  rr <- as.data.frame(roc(data)); roc_col <- intersect(c('roc', 'ROC'), names(rr))[1]",
            "  if (is.na(roc_col)) return(data.frame(id = unique(data$id)))",
            "  stats::aggregate(rr[[roc_col]], list(id = rr$id), function(x) mean(abs(x), na.rm = TRUE))",
            "}",
            "metric_frames <- list(",
            "safe_metric('Active_percent', quote(active_percent(d, tz = 'UTC'))),",
            "safe_metric('ADRR', quote(adrr(d))), safe_metric('AUC', quote(auc(d))),",
            "safe_metric('COGI', quote(cogi(d))), safe_metric('CONGA', quote(conga(d))),",
            "safe_metric('CV', quote(cv_glu(d))), safe_metric('CV_subtypes', quote(cv_measures(d))),",
            "safe_metric('eA1c', quote(ea1c(d))), safe_metric('GMI', quote(gmi(d))),",
            "safe_metric('GRADE', quote(grade(d))), safe_metric('GRADEeu', quote(grade_eugly(d))),",
            "safe_metric('GRADEhyper', quote(grade_hyper(d))), safe_metric('GRADEhypo', quote(grade_hypo(d))),",
            "safe_metric('GVP', quote(gvp(d))), safe_metric('HBGI', quote(hbgi(d))),",
            "safe_metric('LBGI', quote(lbgi(d))), safe_metric('Hyper_Index', quote(hyper_index(d))),",
            "safe_metric('Hypo_Index', quote(hypo_index(d))), safe_metric('IGC', quote(igc(d))),",
            "safe_metric('IQR', quote(iqr_glu(d))), safe_metric('J_index', quote(j_index(d))),",
            "safe_metric('MAD', quote(mad_glu(d))), safe_metric('MAG', quote(mag(d))),",
            "safe_metric('MAGE', quote(mage(d))), safe_metric('Mean', quote(mean_glu(d))),",
            "safe_metric('Median', quote(median_glu(d))), safe_metric('MODD', quote(modd(d))),",
            "safe_metric('M_value', quote(m_value(d))), safe_metric('Percent_Above', quote(above_percent(d))),",
            "safe_metric('Percent_Below', quote(below_percent(d))),",
            "safe_metric('Percent_In_Range', quote(in_range_percent(d))),",
            "safe_metric('Quantiles', quote(quantile_glu(d))), safe_metric('Range', quote(range_glu(d))),",
            "safe_metric('ROC_mean_abs', quote(roc_summary(d))), safe_metric('SD_of_ROC', quote(sd_roc(d))),",
            "safe_metric('SD', quote(sd_glu(d))), safe_metric('SD_subtypes', quote(sd_measures(d))))",
            "out <- Reduce(function(x, y) merge(x, y, by = 'id', all = TRUE), metric_frames)",
            "write.csv(out, args[2], row.names = FALSE, na = '')",
        ]), encoding="utf-8")
        completed = subprocess.run([rscript, "--vanilla", str(script_path), str(input_path), str(output_path)], capture_output=True, text=True, timeout=180)
        if completed.returncode != 0 or not output_path.exists():
            error = (completed.stderr or completed.stdout or "iglu failed").strip()[:500]
            return pd.DataFrame({"visit": visits_lower, "iglu_status": [error] * len(VISITS)})
        iglu_df = pd.read_csv(output_path, encoding="utf-8")
    finally:
        for path in (input_path, output_path, script_path):
            try:
                if path.exists():
                    path.unlink()
            except OSError:
                pass
    if "id" in iglu_df.columns:
        iglu_df = iglu_df.rename(columns={"id": "visit"})
    iglu_df["visit"] = iglu_df["visit"].astype(str).str.lower()
    for visit in visits_lower:
        if visit not in set(iglu_df["visit"]):
            iglu_df = pd.concat([iglu_df, pd.DataFrame([{"visit": visit, "iglu_status": "no glucose data"}])], ignore_index=True)
    return iglu_df


def write_cgm_workbook(patient, visit_frames, output_path):
    has_any = any(not frame.empty for frame in visit_frames.values())
    if not has_any:
        return None
    summary = python_summary(patient, visit_frames)
    iglu = run_iglu_summary(visit_frames)
    if not iglu.empty:
        summary = summary.merge(iglu, on="visit", how="left")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for visit in VISITS:
            visit_frames.get(visit, pd.DataFrame()).to_excel(writer, sheet_name=visit.lower(), index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
    return output_path


def parse_args():
    parser = argparse.ArgumentParser(description="Export Libre CSV files to CGM XLSX workbooks with optional iglu metrics.")
    parser.add_argument("--libre-dir", type=Path, default=DEFAULT_LIBRE_DIR, help="Libre workspace containing data_visits_protected.zip and patient folders.")
    parser.add_argument("--patient", action="append", help="Patient folder to export, e.g. G0001. Can be repeated. Default: all patients with CSV files.")
    parser.add_argument("--output-dir", type=Path, default=None, help="Output directory. Default: <libre-dir>/CGM_EXPORT.")
    parser.add_argument("--no-password", action="store_true", help="Read a local plain data_visits.xlsx. Intended only for trusted local maintenance.")
    return parser.parse_args()


def main():
    args = parse_args()
    libre_dir = args.libre_dir.expanduser().resolve()
    patients = load_patients(libre_dir, no_password=args.no_password)
    wanted = {patient_folder_name(item).upper() for item in args.patient} if args.patient else None
    output_root = (args.output_dir.expanduser().resolve() if args.output_dir else libre_dir / "CGM_EXPORT")
    exported = []
    skipped = []
    for patient in patients:
        if wanted and patient["folder"].upper() not in wanted:
            continue
        if not patient_source_csvs(libre_dir, patient["folder"]):
            skipped.append((patient["folder"], "no csv files"))
            continue
        visit_frames = visit_frames_for_patient(libre_dir, patient)
        output_path = output_root / patient["folder"] / f"{patient['folder']}_CGM.xlsx"
        saved = write_cgm_workbook(patient, visit_frames, output_path)
        if saved:
            exported.append(saved)
            print(f"[OK] {patient['folder']} -> {saved}")
        else:
            skipped.append((patient["folder"], "no glucose data in visit windows"))
    print(f"Exported: {len(exported)}")
    if skipped:
        print(f"Skipped: {len(skipped)}")
        for folder, reason in skipped[:20]:
            print(f"  {folder}: {reason}")


if __name__ == "__main__":
    main()
